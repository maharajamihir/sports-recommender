from openpyxl import load_workbook
import math

wb = load_workbook(filename='./Toughest Sport by Skill.xlsx')
sheet_ranges = wb['Sheet1']
max_dist = math.sqrt(1200)


def getValue(cell):
    return sheet_ranges[cell].value


def iter_rows():
    for row in sheet_ranges.iter_rows():
        yield [cell.value for cell in row]


def get_row(row):
    return [cell.value for cell in sheet_ranges[row]]


def get_dist(input, datapoint):
    if(len(input) != len(datapoint)):
        return RuntimeError("Wrong array length!")
    power = 0
    for i in range(1, len(input)-2):
        power += ((input[i] - datapoint[i]) ** 2)
    return math.sqrt(power)


def get_match(input, datapoint):
    dist = get_dist(input, datapoint)
    dist = dist/max_dist
    dist = dist * 100
    return 100 - dist

def sort(arr):
    return 0


def recommend(input):
    output = [[0 for x in range(2)] for y in iter_rows()]
    for row in range(0, len(output)-2):
        output[row][0] = get_row('A')[row+1]
        output[row][1] = get_match(input, get_row(row+2))

    return output


print(max_dist)
print(get_row('1'))
print(get_match([0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
      get_row('2')))
print(get_dist([0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
      get_row('2')))
print(recommend([0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]))
